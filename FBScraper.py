#!usr/bin/env python
# Python Program to mine data from a Brand's Facebook Account
__author__ = "Nihinlolamiwa Fajemilehin, Timothy Shirgba, Sunnny Shokeen"
__copyright__ = "Copyright 2019, KITC"
__version__ = "1.0.2"

import json
import facebook
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.parse import urlencode
from urllib.request import urlopen
import os.path

filename_root = os.path.abspath(os.path.dirname(__file__))

# Function that gets all sustainability keywords from an excel file and returns in a list
def getSustainabilityKeywords(filename):

    # open the keyword excel workbook
    kw_wb_input = load_workbook(filename=filename)

    # select SusMon_Keywords.xlsx
    kw_sheet = kw_wb_input.active

    # Get all the keywords from the first column
    m_row = kw_sheet.max_row

    # Initialize keyword list
    keywordList = []

    for i in range(1, m_row + 1):
        cell_obj = kw_sheet.cell(row=i, column=1)
        keywordList.append(cell_obj.value)

    return keywordList

# Function that gets Access Token and Page ID from an Excel file
def getInputData(filename):

    # open the excel input workbook
    wb_input = load_workbook(filename=filename, read_only=True)

    # select FBPage_Input.xlsx
    sheet = wb_input.active

    # get Page ID
    pageIdLabel = sheet['B1']
    pageIdValue = sheet['B2']
    pageIdValue = round(pageIdValue.value)

    # get Brand name
    pageNameLabel = sheet['C1']
    pageNameValue = sheet['C2']
    pageNameValue = pageNameValue.value

    # get Access Token
    pageAccessTokenLabel = sheet['D1']
    pageAccessTokenValue = sheet['D2']
    pageAccessTokenValue = pageAccessTokenValue.value

    # Store ID and Access token in a dictionary
    d = dict()
    d['brands'] = [{'id':pageIdValue}] #at the moment we are only expecting a single brand, but the rest of the code expects multiuple brands
    d['accessToken'] = pageAccessTokenValue
    return d

# Function that calls Input Data and Keywords functions 
def loadConfig():
  
  input_filename = filename_root + '/FBPage_Input.xlsx'
  keyword_filename = filename_root + '/SusMon_Keywords.xlsx'

  config = getInputData(input_filename)
  config['keywords'] = getSustainabilityKeywords(keyword_filename)

  return config

# Function that calls Facebook's Graph API to grab all the posts for a particular brand page
def getPostsForBrand(brandID, token):
  uri="https://graph.facebook.com/v4.0/{}/posts/".format(str(brandID))
  fields = "id,created_time,message,shares.summary(true).limit(0),comments,likes.summary(true),reactions.type(LOVE).limit(0).summary(total_count).as(Love),reactions.type(WOW).limit(0).summary(total_count).as(Wow),reactions.type(HAHA).limit(0).summary(total_count).as(Haha),reactions.type(SAD).limit(0).summary(1).as(Sad),reactions.type(ANGRY).limit(0).summary(1).as(Angry)"
  url= "{uri}?fields={fields}&access_token={token}&limit={limit}"\
    .format(**{'uri':uri,'fields':fields,'token': str(token), 'limit':50 })
    
  #url = "https://graph.facebook.com/v4.0/{pageID}/posts/?fields=id,created_time,message,shares.summary(true).limit(0),comments.summary(true).limit(0),likes.summary(true),reactions.type(LOVE).limit(0).summary(total_count).as(Love),reactions.type(WOW).limit(0).summary(total_count).as(Wow),reactions.type(HAHA).limit(0).summary(total_count).as(Haha),reactions.type(SAD).limit(0).summary(1).as(Sad),reactions.type(ANGRY).limit(0).summary(1).as(Angry)&access_token={token}&limit=50".format(**{'pageID': str(pageId), 'token': str(token)})
  try:
    facebook_connection = urlopen(url)
    data = facebook_connection.read().decode('utf8')
    json_object = json.loads(data)
  except Exception as ex:
    print(ex)
  return json_object["data"]

# Check to see if the post is relevant
def isRelevant(post,keywords):
  
  for keyword in keywords:
    
    if 'message' in post and keyword in post['message']:
      return True

  return False

# Exports post data to a csv
def exportResultsToSCV(results):
  wb = Workbook()

  # List for all post wtih comments
  postWithComments = []

  # set file path
  filePath = filename_root+'/FBOutput.xlsx'

  # Save file in the path
  wb.save(filePath)

  # Load workbook
  wb = load_workbook(filePath)

  sheet = wb.active

  # Create Headers in created output file in a row
  sheet['A1'] = 'Brand ID'
  sheet['B1'] = 'Post ID'
  sheet['C1'] = 'Created Time'
  sheet['D1'] = '#Comments'
  sheet['E1'] = '#like'
  sheet['F1'] = '#love'
  sheet['G1'] = '#wow'
  sheet['H1'] = '#haha'
  sheet['I1'] = '#sad'
  sheet['J1'] = '#angry'
  sheet['K1'] = 'message'
  sheet['L1'] = 'Shares'

  # Save data retreived from Facebook in next row
  row = 2

  for brand in results:
    
    brandID = brand['brandId']

    for post in brand['posts']:
      sheet['A'+str(row)] = brandID
      sheet['B'+str(row)] = post['id']
      sheet['C'+str(row)] = post['created_time']
      if 'comments' in post:
        sheet['D'+str(row)] = len(post['comments']['data'])
      else:
        sheet['D'+str(row)] = '0'
      sheet['E'+str(row)] = len(post['likes']['data'])
      sheet['F'+str(row)] = post['Love']['summary']['total_count']
      sheet['G'+str(row)] = post['Wow']['summary']['total_count']
      sheet['H'+str(row)] = post['Haha']['summary']['total_count']
      sheet['I'+str(row)] = post['Sad']['summary']['total_count']
      sheet['J'+str(row)] = post['Angry']['summary']['total_count']
      sheet['K'+str(row)] = post['message']
      if 'shares' in post:
        sheet['L'+str(row)] = post['shares']['count']
      else:
        sheet['L'+str(row)] = '0'
      if 'comments' in post:
      
       for comment in post['comments']['data']:
          # Call function that stores post comments in a different Excel file 
          postWithComments.append(comment)          
      row = row + 1 

  storePostComments(postWithComments)

  wb.save(filePath)

# Function that store Post comments in an excel file
def storePostComments(commentDetails):

  wb = Workbook()

  # set file path
  filePath = filename_root+'/PostComments.xlsx'

  # Save file in the path
  wb.save(filePath)

  # Load workbook
  wb = load_workbook(filePath)

  sheet = wb.active

  # Create Headers in created output file in a row
  sheet['A1'] = 'Post ID'
  sheet['B1'] = 'Comment Datetime'
  sheet['C1'] = 'Comment Content'

  row = 2
  for commentDetail in commentDetails:
    sheet['A'+str(row)] = commentDetail['id']
    sheet['B'+str(row)] = commentDetail['created_time']
    sheet['C'+str(row)] = commentDetail['message']
    row += 1

  wb.save(filePath)

# Function that is called first when the program is executed
def main():
  config = loadConfig()
  
  results = []
  
  for brand in config['brands']:

    relevantPosts = []
    posts = getPostsForBrand(brand['id'], config['accessToken'])

    for post in posts:
      # print(post)
      if isRelevant(post, config['keywords']):
        relevantPosts.append(post)

    results.append({'brandId':brand['id'], 'posts':relevantPosts})

  exportResultsToSCV(results)

if __name__ == '__main__':
    main()