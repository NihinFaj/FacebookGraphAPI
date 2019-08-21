#!usr/bin/env python
# Python Program to mine data from a Brand's Facebook Account
import json
import facebook
# The library needed to read an excel file using Python
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.parse import urlencode
import pandas as pd
from urllib.request import urlopen
import numpy as np
import math

# The location of the input file
filename_root = 'C:/Users/Synergos/OneDrive/Documents/SusMon_Facebook_API'
input_filename = filename_root + '/FBPage_Input.xlsx'
keyword_filename = filename_root + '/SusMon_Keywords.xlsx'
token = ''

# Function that reads Access Token and Page ID from an Excel file
def getInputData():

    # open the excel input workbook
    wb_input = load_workbook(filename=input_filename, read_only=True)

    # select FBPage_Input.xlsx
    sheet = wb_input.active

    # get Page ID
    pageIdLabel = sheet['B1']
    pageIdValue = sheet['B2']
    pageIdValue = round(pageIdValue.value)

    # get Brand name
    pageNameLabel = sheet['C1']
    pageNameValue = sheet['C2']
    pageNameValue = pageNameValue.value

    # get Access Token
    pageAccessTokenLabel = sheet['D1']
    pageAccessTokenValue = sheet['D2']
    pageAccessTokenValue = pageAccessTokenValue.value

    # Store ID and Access token in a dictionary
    d = dict()
    d['pageId'] = pageIdValue
    d['accessToken'] = pageAccessTokenValue
    return d

# Function to get all sustainability keywords from an excel file and return in a list
def getSustainabilityKeywords():

    # open the keyword excel workbook
    kw_wb_input = load_workbook(filename=keyword_filename)

    # select SusMon_Keywords.xlsx
    kw_sheet = kw_wb_input.active

    # Get all the keywords from the first column
    m_row = kw_sheet.max_row

    # Initialize keyword list
    keywordList = []

    for i in range(1, m_row + 1):
        cell_obj = kw_sheet.cell(row=i, column=1)
        keywordList.append(cell_obj.value)

    return keywordList

# Function that creates output file and connects to Facebook afterwards
def createOutputFile():

    responseFromFacebook = connectToFacebook()

    wb = Workbook()

    # set file path
    filePath = 'C:/Users/Synergos/OneDrive/Documents/SusMon_Facebook_API/FBOutput.xlsx'

    # Save file in the path
    wb.save(filePath)

    # Load workbook
    wb = load_workbook(filePath)

    sheet = wb.active

    # Create Headers in created output file in a row
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Fan Count'
    sheet['C1'] = 'Posts'
    sheet['D1'] = 'Link'

    # Save data retreived from Facebook in next row
    sheet['A2'] = responseFromFacebook.get("name")
    sheet['B2'] = responseFromFacebook.get("fan_count")
    sheet['C2'] = json.dumps(responseFromFacebook.get("posts"))
    sheet['D2'] = responseFromFacebook.get("link")

    wb.save(filePath)


# Function to initialize and connect to FacebookGraphAPI
def connectToFacebook():
    # Call function to get input data
    values = getInputData()

    pageId = values.get("pageId")
    # accessToken = values.get("accessToken")

    token = values.get("accessToken")
    # graph = facebook.GraphAPI(token)
    # fields = ['first_name', 'location{location}','email','link']
    # profile = graph.get_object(pageId, fields='name,fan_count,posts,link')
    # return desired fields
    # print(json.dumps(profile, indent=4))

    url = "https://graph.facebook.com/v4.0/"+str(pageId)+"/posts/?fields=id,created_time,message,shares.summary(true).limit(0),comments.summary(true).limit(0),likes.summary(true),reactions.type(LOVE).limit(0).summary(total_count).as(Love),reactions.type(WOW).limit(0).summary(total_count).as(Wow),reactions.type(HAHA).limit(0).summary(total_count).as(Haha),reactions.type(SAD).limit(0).summary(1).as(Sad),reactions.type(ANGRY).limit(0).summary(1).as(Angry)&access_token="+str(token)+"&limit=50"
    try:
        facebook_connection = urlopen(url)
        data = facebook_connection.read().decode('utf8')
        json_object = json.loads(data)
        posts = json_object["data"]
        df = pd.DataFrame(posts)

    except Exception as ex:
        print(ex)

    return df

# Function to scan each post message for sustainability keywords
def checkPostForKeywords():

    keywordsList = getSustainabilityKeywords()

    facebookPosts = connectToFacebook()
    df = facebookPosts

    # Get Total length of keyword list
    totalNumOfKeywords = len(keywordsList)

    # loop through all rows of the in the dataframe
    for index, row in df.iterrows():

        individualMsg = row['message']
        individualId = row['id']

        # Check if message is Nan
        if pd.isnull(individualMsg):
            continue

        # If not Nan encode and search against the list of keywords
        tw_txtu = individualMsg.encode('utf-8')

        term_count = 0
        while term_count < totalNumOfKeywords:

            # loop through our relevant terms
            termu = keywordsList[term_count].encode('utf-8')
            if termu in tw_txtu:
                # Call function that calls Facebook API with Post ID and save retrieved data into Output Excel file
                callFacebookAndSaveData(individualId)
                break
            else:
                term_count += 1

# Function that calls Facebook API with ID and save retrieved data into Output Excel file
def callFacebookAndSaveData(individualId):

    print("ID is " + individualId)
    # Call function to get input data
    values = getInputData()
    token = values.get("accessToken")
    graph = facebook.GraphAPI(token)

    url = "https://graph.facebook.com/v4.0/"+ individualId + "/?fields=id,created_time,message,shares.summary(true).limit(0),comments.summary(true).limit(0),likes.summary(true),reactions.type(LOVE).limit(0).summary(total_count).as(Love),reactions.type(WOW).limit(0).summary(total_count).as(Wow),reactions.type(HAHA).limit(0).summary(total_count).as(Haha),reactions.type(SAD).limit(0).summary(1).as(Sad),reactions.type(ANGRY).limit(0).summary(1).as(Angry)&access_token=" + token + "&limit=50"

    try:
        facebook_connection = urlopen(url)
        data = facebook_connection.read().decode('utf8')
        json_object = json.loads(data)
        # posts = json_object["data"]
        df = pd.DataFrame(json_object)
        print(df)
        
        df['Angry'] = df['Angry'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['Angry'] = df['Angry'].str.replace(',(.*?)}}', '')
        df['Haha'] = df['Haha'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['Haha'] = df['Haha'].str.replace('}}', '')
        df['Love'] = df['Love'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['Love'] = df['Love'].str.replace('}}', '')
        df['Sad'] = df['Sad'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['Sad'] = df['Sad'].str.replace(',(.*?)}}', '')
        df['Wow'] = df['Wow'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['Wow'] = df['Wow'].str.replace('}}', '')
        df['comments'] = df['comments'].astype(str).str.replace('{\'data\':(.*?)count\': ', '')
        df['comments'] = df['comments'].str.replace(',(.*?)}}', '')
        df['likes'] = df['likes'].astype(str).str.replace('{\'(.*?)count\':', '')
        df['likes'] = df['likes'].str.replace(',(.*?)}}', '')
        df['shares'] = df['shares'].astype(str).str.replace('{\'count\': ', '')
        df['shares'] = df['shares'].str.replace('}', '')
        df['date'], df['time'] = df['created_time'].astype(str).str.split('T', 1).str
        df['time'] = df['time'].str.replace('[+]0000', '')
        df.to_csv("Facebook Posts.csv") 
        
        print("The END")
    except Exception as ex:
        print(ex)

    

def main():

    # Call function to import all keywords
    # getSustainabilityKeywords()

    # Call createOutputFile function
    # createOutputFile()

    # connectToFacebook()
    checkPostForKeywords()

if __name__ == '__main__':
    main()