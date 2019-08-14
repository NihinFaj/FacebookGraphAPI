

Skip to content
Using Kent IT Consultancy Mail with screen readers
python code 

3 of 3
Fwd: Example Python Code
Inbox
x

Sunny Shokeen
Attachments
Wed, 17 Jul, 11:28
to Timothy, A, Kwabena, Mai, me, Jason

Hi Team,

Great news, Philip has shared the code with us. Please have look and familiarize yourselves with it, keeping into consideration that it is just an example and should not be considered as a template for what the client is expecting as mentioned by Philip in his email.

---------- Forwarded message ---------
From: Philip Dunn <philip@susmon.com>
Date: Wed, 17 Jul 2019 at 09:59
Subject: Example Python Code
To: Sunny Shokeen <ss995@kitc-solutions.co.uk>


Good Morning Sunny,

It was good to meet you and the rest of the KITC team on Thursday. I picked up a couple of actions, one to provide you with some example code for the Twitter application we have already written and the other to create and provide you access to a Susmon Facebook Developer account.

I’m fulfilling the first of those actions here by sending you some example code. Python is an Object Oriented programming language, but my programming experience is primarily procedural (C, Fortran, Pascal etc.) . You will see from the way the code is written it does not make much use of objects except from predefined libraries. So please don’t take this as a template for how I’d like to see any POC coding done – I’m sure you all have better ideas. Also attached is the 5 lines of code I wrote just to quickly test access to the Graph API from Python.

I’ve yet to set up the Susmon Facebook Developer account – I need some input from one of our team who is on holiday at the moment. I should get that done next week.

In the meantime, if you have any other questions do not hesitate to get in touch.

Best wishes,

Philip

 

Philip Dunn

+44 (0)7598 110189

philip@susmon.com

signature_1755007581

www.susmon.com

Sustainability Monitor Ltd · Registered in England & Wales · Company No. 10586487
Registered Office: Innovation Centre · University Road · Canterbury · Kent CT2 7FG · United Kingdom

This message is private and confidential. If you have received this message in error, please notify us and remove it from your system.



Many Thanks, 

3 Attachments

Sunny Shokeen
Attachments
Wed, 17 Jul, 16:09
to Timothy, Kwabena, A, Mai, me, Jason

Hello everyone,

Philip has shared the social media process slides with us. Please have a look at it. Hard Copy of Suswatch is with Tim.

---------- Forwarded message ---------
From: Philip Dunn <philip@susmon.com>
Date: Wed, 17 Jul 2019 at 15:56
Subject: Re: Example Python Code
To: Sunny Shokeen <ss995@kitc-solutions.co.uk>


Hi Sunny,

Happy to share the Social Media Process slides which I have attached.

SusWatch is one of the key research products we are developing and therefore I do not want to share electronic copies of that at the moment. If there are particular slides that would be of use to you, let me know and I may be able to share some limited content.

I’ll come back to you on dates for meetings tomorrow – I know I have some conflicts and I’m just checking with Ronnie whether the alternative dates I have will work for him.

Best wishes,

Philip

 

 

 

Philip Dunn

+44 (0)7598 110189

philip@susmon.com

signature_1755007581

www.susmon.com

Sustainability Monitor Ltd · Registered in England & Wales · Company No. 10586487
Registered Office: Innovation Centre · University Road · Canterbury · Kent CT2 7FG · United Kingdom

This message is private and confidential. If you have received this message in error, please notify us and remove it from your system.

 

From: Sunny Shokeen <ss995@kitc-solutions.co.uk>
Date: Wednesday, 17 July 2019 at 11:49
To: Philip Dunn <philip@susmon.com>
Subject: Re: Example Python Code

 

Good Morning Philip, 

 

It was great meeting you as well and thanks for sharing the code with us. Our team is currently working on design specifications and success criteria which we will validate with you during our discovery playback (invite for which I have sent in my previous email). Could you please share the presentation slides for Social Media Process and Suswatch, we have the hard copy but if we could get the soft copy as well it would be really helpful. 

I will keep you informed about the process and get back to you if we have any questions.

 

On Wed, 17 Jul 2019 at 09:59, Philip Dunn <philip@susmon.com> wrote:

Good Morning Sunny,

It was good to meet you and the rest of the KITC team on Thursday. I picked up a couple of actions, one to provide you with some example code for the Twitter application we have already written and the other to create and provide you access to a Susmon Facebook Developer account.

I’m fulfilling the first of those actions here by sending you some example code. Python is an Object Oriented programming language, but my programming experience is primarily procedural (C, Fortran, Pascal etc.) . You will see from the way the code is written it does not make much use of objects except from predefined libraries. So please don’t take this as a template for how I’d like to see any POC coding done – I’m sure you all have better ideas. Also attached is the 5 lines of code I wrote just to quickly test access to the Graph API from Python.

I’ve yet to set up the Susmon Facebook Developer account – I need some input from one of our team who is on holiday at the moment. I should get that done next week.

In the meantime, if you have any other questions do not hesitate to get in touch.

Best wishes,

Philip

 

Philip Dunn

+44 (0)7598 110189

philip@susmon.com

Error! Filename not specified.

www.susmon.com

Sustainability Monitor Ltd · Registered in England & Wales · Company No. 10586487
Registered Office: Innovation Centre · University Road · Canterbury · Kent CT2 7FG · United Kingdom

This message is private and confidential. If you have received this message in error, please notify us and remove it from your system.



 

Many Thanks, 

Sunny Shokeen | Student Consultant

www.kitc-solutions.co.uk
CONFIDENTIALITY NOTICE:This message (including any attachments) may contain confidential, proprietary, privileged and/or private information. The information is intended to be for the use of the individual or entity designated above. If you are not the intended recipient of this message, please notify the sender immediately, and delete the message and any attachments. Any disclosure, reproduction, distribution or other use of this message or any attachments by an individual or entity other than the intended recipient is prohibited.



Many Thanks

Attachments area
Thanks a lot for sharing.Sure, will do that.I haven't received it yet.

# prototype program to read tweets from companies of interest for the last week and save those tweets that relate to sustainability.
# it then searches for replies to those companies and matches them to tweets and updates the reply count and saves reples in a separate sheet per co
# author: Philip Dunn
# date: 14/04/2019
#
# Change log:
# date: 08/05/2019 deal with brands in the input file.Brands are listed with the tag "brand" in the rows immediately following the paraent company in the input file.
# date: 22/05/2019 deal with month start/end happening in the period being collected
# date: 24/06/2019 get rid of redundant Sheet1 in output workbook
# date: 15/07/2019 update comments to make clearer
import tweepy
import datetime
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from susmon_weekly_replies import replies_to_co

def initialise_sheet(target_ws):
	# insert header row in a worksheet
	c=target_ws.cell(row=1,column=1,value="Tweet Identifier")
	c=target_ws.cell(row=1,column=2,value="User Name")
	c=target_ws.cell(row=1,column=3,value="Timestamp")
	c=target_ws.cell(row=1,column=4,value="Keyword")
	c=target_ws.cell(row=1,column=5,value="Tweet Text")
	c=target_ws.cell(row=1,column=6,value="Replies")
	c=target_ws.cell(row=1,column=7,value="Retweets")
	c=target_ws.cell(row=1,column=8,value="Likes")
	c=target_ws.cell(row=1,column=9,value="Followers")
	c=target_ws.cell(row=1,column=10,value="URL")

def process_tweet(status,target_ws,terms,num_terms,args):
	#process individual tweets identifying relevance and saving key data to an Excel worksheet
	timestamp=status.created_at
	if hasattr(status,'retweeted_status')==False:
		if status.in_reply_to_user_id is None:
			#this is not a reply or a retweet - so an original tweet
			args[1]+=1
			user_name=status.user.name
			user_nameu=user_name.encode('utf-8')
			print(user_nameu+" has tweeted something\n")
			try:
				tw_txt=status.extended_tweet["full_text"]
			except AttributeError:
				tw_txt=status.full_text
			tw_txtu=tw_txt.encode('utf-8')
			#we only want to save tweets that have the terms in which we are interested
			term_count=0
			while term_count<num_terms:
				#loop through our relevant terms
				termu=terms[term_count].encode('utf-8')
				if termu in tw_txtu:
					#save the tweet as soon as we find a match and exit loop
					c=target_ws.cell(row=args[2],column=1,value=status.id_str)
					c=target_ws.cell(row=args[2],column=2,value=user_nameu)
					c=target_ws.cell(row=args[2],column=3,value=timestamp.strftime("%d/%m/%y %H:%M:%S"))
					c=target_ws.cell(row=args[2],column=4,value=termu)
					c=target_ws.cell(row=args[2],column=5,value=tw_txtu)
					c=target_ws.cell(row=args[2],column=6,value=0)
					c=target_ws.cell(row=args[2],column=7,value=status.retweet_count)
					c=target_ws.cell(row=args[2],column=8,value=status.favorite_count)
					c=target_ws.cell(row=args[2],column=9,value=status.user.followers_count)
					url="https://twitter.com/i/web/status/"+status.id_str
					c=target_ws.cell(row=args[2],column=10,value=url)
					args[2]+=1
					print(user_nameu+" has tweeted something relevant\n")
					break
				else:
					term_count+=1
		else:
			user_name=status.user.name
			user_nameu=user_name.encode('utf-8')
			print(user_nameu+" has replied to something\n")
	else:
		user_name=status.user.name
		user_nameu=user_name.encode('utf-8')
		print(user_nameu+" has retweeted something\n")

	args[0]=timestamp
	return args

def main():
	#Consumer and Access tokens, these are unique to each approved application / project and are private 
	consumer_key=""
	consumer_secret=""
	access_token=""
	access_token_secret=""

	relevant_terms=[]
	companies=[]
	company_names=[]
	brands=[]
	brand_names=[]
	num_terms=0
	num_cos=0
	co_count=0
	latest_tweet=datetime.datetime.today()
	earliest_tweet=latest_tweet-datetime.timedelta(days=7)
	periods=[]
	num_periods=1
	period_count=1
	#arguments list passed and updated when processing a tweet 0-timestamp, 1-tweet count, 2-relevant tweet count & row index
	arg_list=[latest_tweet,0,0]
	filename_root="/home/18tcdata/Susdev/"
	input_filename=filename_root+"susmon_inputs_new2.xlsx"

	#do some date range (period) processing to deal with month boundaries. If we have gone over a month boundary and it's not the first of the month,
	#we will want to create two files one for end of previous month and one for start of current month
	if earliest_tweet.month!=latest_tweet.month and latest_tweet.day!=1:
		month_start=latest_tweet.replace(day=1)
		periods.append([earliest_tweet, month_start])
		periods.append([month_start,latest_tweet])
		num_periods=2
	else:
		periods.append([earliest_tweet,latest_tweet])

	#open our input workbook
	wb_input=load_workbook(filename=input_filename,read_only=True)

	#get list of companies - parents and brands
	ws_companies=wb_input['CoTwitterIDs']
	print("Following: ")
	rows=2
	for row in ws_companies.rows:
		c=ws_companies.cell(row=rows,column=1)
		if c.value!="brand":
			c=ws_companies.cell(row=rows,column=4)
			print(c.value)
			company_names.append(c.value)
			c=ws_companies.cell(row=rows,column=3)
			companies.append(c.value)
			brands.append([])
			brand_names.append([])
			co_count+=1
		elif c.value=="brand":
			c=ws_companies.cell(row=rows,column=3)
			brands[co_count-1].append(c.value)
			c=ws_companies.cell(row=rows,column=4)
			print(c.value)
			brand_names[co_count-1].append(c.value)
		else:
			print("Invalid cell value\n")
			break
		rows+=1
	#inputs has a header row so need to reduce count by 1
	num_cos=co_count-1

	#get list of relevant search terms
	ws_terms=wb_input['SusmonTerms']
	print("Searching for: ")
	rows=1
	for row in ws_terms.rows:
		c=ws_terms.cell(row=rows,column=1)
		relevant_terms.append(c.value)
		print(c.value)
		rows+=1
	num_terms=len(relevant_terms)

	# close our inputs workbook
	wb_input.close()

	#connect to the API
	auth=tweepy.OAuthHandler(consumer_key, consumer_secret)
	auth.set_access_token(access_token, access_token_secret)
	api=tweepy.API(auth)

	#we may be creating two output files if we are going over a month end - defined by period
	for period in periods:
		#here we'll need to create the output file, and loop through the companies, creating a sheet and then
		#getting tweets and processing them
		wb_output=Workbook()
		ws1=wb_output.active
		ws1.title="Summary"
		#create a header row for summary sheet
		c=ws1.cell(row=1,column=1,value="User Id")
		c=ws1.cell(row=1,column=2,value="Original Tweet Count")
		c=ws1.cell(row=1,column=3,value="Relevant Tweet Count")
		c=ws1.cell(row=1,column=4,value="Tweets @")

		co_count=0
		while co_count<num_cos:
			#create and initialise a worksheet for this company's tweets
			ws=wb_output.create_sheet(company_names[co_count])
			initialise_sheet(ws)
			arg_list[1]=0
			arg_list[2]=2
			#get tweets for parent company
			for tweet in tweepy.Cursor(api.user_timeline,user_id=companies[co_count],tweet_mode='extended').items(400):
				if tweet.created_at<period[1] and tweet.created_at>=period[0]:
					arg_list=process_tweet(tweet,ws,relevant_terms,num_terms,arg_list)
				elif tweet.created_at<period[0]:
					break
				else:
					pass

			#now get tweets for each of the brands which we'll just add to the parent company worksheet
			for brand in brands[co_count]:
				for tweet in tweepy.Cursor(api.user_timeline,user_id=brand,tweet_mode='extended').items(400):
					if tweet.created_at<period[1] and tweet.created_at>=period[0]:
						arg_list=process_tweet(tweet,ws,relevant_terms,num_terms,arg_list)
					elif tweet.created_at<period[0]:
						break
					else:
						pass

			#write the parent company name, total number of tweets and number of relevant tweets to Summary worksheet
			c=ws1.cell(row=co_count+2,column=1,value=company_names[co_count])
			c=ws1.cell(row=co_count+2,column=2,value=arg_list[1])
			c=ws1.cell(row=co_count+2,column=3,value=arg_list[2]-2)
			co_count+=1

		#generate the filename an save the file
		date_for_name=period[1]-datetime.timedelta(days=1)
		output_filename=filename_root+"susmon_weekly_"+date_for_name.strftime("%y%m%d")+".xlsx"
		wb_output.save(output_filename)

		#now go back through the file and check for replies to any of our relevant tweets
		co_count=0
		while co_count<num_cos:
			to_count=replies_to_co(api,wb_output,output_filename,company_names[co_count],brand_names[co_count],period[1])
			c=ws1.cell(row=co_count+2,column=4,value=to_count)
			co_count+=1

		#save the file again with any updated replies data
		wb_output.save(output_filename)
		wb_output.close()

		if num_periods==2 and period_count==1:
			#if we are going to have to do this again, let's wait another 15 minutes to avoid being rate limited by the API
			period_count+=1
			print("Going to sleep for 15 minutes\n")
			time.sleep(1000)
		
		
	print("that's all the tweets for now")

if __name__=='__main__':
	main()
susmon_weekly_anon.py
Displaying first_FB_anon.py.